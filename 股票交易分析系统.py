"""
股票交易盈亏分析系统
功能：自动分析股票交易记录，计算盈亏并生成报告
支持三种输入：
  1. Excel文件（券商导出）— 文件名格式：YYYY-MM-DD-两融-当日成交汇总.xlsx
  2. 图片文件（手机App截图）— 文件名格式：YYYY-MM-DD-手机交易.png/.jpg/.jpeg
  3. 图片文件（平安证券截图）— 文件名格式：YYYYMMDD_平安.png
作者：WorkBuddy
版本：v4.5
更新日期：2026-05-10

使用说明：
1. 将券商导出的Excel交易记录文件、手机App截图或平安证券截图放到当前文件夹
2. 运行此脚本，自动处理所有未归档的文件
3. 处理后的原始文件会自动归档到history文件夹
4. HTML报告生成到reports文件夹（单日报告 + 汇总可视化报告）
5. Excel汇总文件支持去重更新（同日期+同来源的数据会覆盖）
6. 汇总可视化报告支持时间筛选和多级数据钻取
7. 同一天可以同时有Excel、手机截图和平安截图多种输入，数据会自动合并
8. v4.3新增：佣金（万一/双向/最低5元）和印花税（万五/卖出单边）计算
9. v4.4新增：个股累计盈亏总览标签页、盈亏日历热力图（含连续亏损预警）
10. v4.5新增：跨账户同股配对、单边交易标记（未平仓）、连续亏损/回撤统计

文件结构：
- 当前文件夹：待处理的Excel文件和图片文件
- reports/：HTML报告文件（单日报告 + 汇总可视化报告）
- history/：已处理的原始文件
- 股票交易盈亏汇总.xlsx：累计盈亏数据（去重更新）
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import shutil
import glob
import re

# ==================== 配置参数 ====================
EXCEL_OUTPUT = '股票交易盈亏汇总.xlsx'  # Excel汇总文件
REPORTS_DIR = 'reports'                  # HTML报告文件夹
HISTORY_DIR = 'history'                  # 原始文件归档文件夹
TEMPLATES_DIR = os.path.join(REPORTS_DIR, 'templates')  # HTML模板文件夹
SUMMARY_TEMPLATE = os.path.join(TEMPLATES_DIR, 'summary_report.html')  # 汇总报告模板

# 确保文件夹存在
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(HISTORY_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# 交易成本配置
COMMISSION_RATE = 0.0001      # 佣金费率：万一（0.01%），双向征收
MIN_COMMISSION = 5             # 最低佣金：5元/笔
STAMP_DUTY_RATE = 0.0005      # 印花税费率：万五（0.05%），仅卖出单边征收


# ==================== 核心处理函数 ====================

def extract_date_from_filename(filename):
    """从文件名提取日期，支持 YYYY-MM-DD 和 YYYYMMDD 格式"""
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if date_match:
        return date_match.group(1)
    date_match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if date_match:
        return f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
    return datetime.now().strftime('%Y-%m-%d')


def get_source_from_filename(filename):
    """从文件名判断数据来源"""
    fname_lower = filename.lower()
    if fname_lower.endswith('.xlsx'):
        return '两融账户'
    elif fname_lower.endswith('.xls'):
        if '平安' in fname_lower:
            return '平安账户'
        return '两融账户'
    elif any(fname_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg']):
        if '平安' in fname_lower:
            return '平安账户'
        return '手机账户'
    return '未知'


# ==================== 图片OCR解析 ====================

# OCR Reader单例（避免每次解析图片都重新加载模型，省3-5秒/次）
_OCR_READER = None

def get_ocr_reader():
    """获取easyocr Reader单例，首次调用时加载模型，后续复用"""
    global _OCR_READER
    if _OCR_READER is None:
        import easyocr
        _OCR_READER = easyocr.Reader(['ch_sim', 'en'], gpu=False, verbose=False)
    return _OCR_READER


# OCR名称修正映射表（OCR经常截断/误读股票名称，用此表修正）
STOCK_NAME_CORRECTIONS = {
    '688503': '聚和材料',   # OCR常读成"聚和材"缺"料"
    '688041': '海光信息',
    '300502': '新易盛',
    '688025': '杰普特',
    '300308': '中际旭创',
    '002436': '兴森科技',
    '600498': '烽火通信',
    '688035': '德邦科技',
    '002594': '比亚迪',
    '300750': '宁德时代',
    '300394': '天孚通信',
    '688256': '寒武纪',
    '300014': '亿纬锂能',
    '002156': '通富微电',
    '688062': '迈威生物',
    '300660': '江苏雷利',
    '688778': '厦钨新能',
    '300450': '先导智能',
    '601208': '东材科技',
}


def parse_image_trades(image_path):
    """用手机App截图识别交易记录 - v2.0 改进版
    策略：不按行分组，而是提取所有关键信息后按股票代码聚合
    """
    # 修复Windows下torch/getpass环境变量问题
    import os as _os
    if not _os.environ.get('USERNAME'):
        _os.environ['USERNAME'] = 'Ryan'
    if not _os.environ.get('USER'):
        _os.environ['USER'] = 'Ryan'

    try:
        import easyocr
    except ImportError:
        print(f"[错误] 缺少OCR依赖，请运行：pip install easyocr")
        return pd.DataFrame()

    print(f"  正在识别图片：{os.path.basename(image_path)}")
    reader = get_ocr_reader()
    result = reader.readtext(image_path)

    # 提取所有文本项，保留位置信息
    items = []
    for bbox, text, conf in result:
        y_center = (bbox[0][1] + bbox[2][1]) / 2
        x_center = (bbox[0][0] + bbox[2][0]) / 2
        items.append({'text': text.strip(), 'y': y_center, 'x': x_center, 'conf': conf})

    items.sort(key=lambda r: r['y'])

    # 第一步：按y坐标分组（行间距阈值30px，支持逐笔成交分离）
    rows = []
    current_row = []
    last_y = None
    y_threshold = 30

    for item in items:
        if last_y is None or abs(item['y'] - last_y) <= y_threshold:
            current_row.append(item)
        else:
            if current_row:
                rows.append(current_row)
            current_row = [item]
        last_y = item['y']

    if current_row:
        rows.append(current_row)

    for row in rows:
        row.sort(key=lambda r: r['x'])

    # 第二步：从每行提取关键信息
    raw_records = []
    for row in rows:
        texts = [r['text'] for r in row]
        text_combined = ' '.join(texts)

        # 提取股票代码（6位数字）
        code_match = re.search(r'\b(\d{6})\b', text_combined)
        if not code_match:
            continue
        stock_code = code_match.group(1)

        # 提取方向
        direction = None
        if '买入' in text_combined or '买' in text_combined:
            direction = '证券买入'
        elif '卖出' in text_combined or '卖' in text_combined:
            direction = '证券卖出'
        else:
            continue

        # 提取股票名称：在代码之前或之后的中文
        stock_name = ''
        code_idx = text_combined.find(stock_code)
        
        # 尝试代码前面
        if code_idx > 0:
            prefix = text_combined[:code_idx].strip()
            prefix = re.sub(r'\d{1,2}:\d{2}:\d{2}', '', prefix)
            prefix = re.sub(r'买入|卖出|买|卖|成交', '', prefix)
            prefix = prefix.strip()
            cn_match = re.findall(r'[\u4e00-\u9fff]+', prefix)
            if cn_match:
                stock_name = cn_match[-1]
        
        # 如果前面没找到，尝试代码后面
        if not stock_name:
            suffix = text_combined[code_idx + 6:].strip()  # 代码后6位之后
            suffix = re.sub(r'\d+\.?\d*', '', suffix)  # 移除数字
            suffix = re.sub(r'买入|卖出|买|卖|成交', '', suffix)
            suffix = suffix.strip()
            cn_match = re.findall(r'[\u4e00-\u9fff]+', suffix)
            if cn_match:
                stock_name = cn_match[0]  # 取第一个

        # 提取数字（价格、金额、数量）
        numbers_text = text_combined.replace(stock_code, '')
        numbers = re.findall(r'\d+\.?\d*', numbers_text)
        numbers = [float(n) for n in numbers if float(n) > 0]

        # 过滤掉时间戳中的数字（格式 HH:MM:SS 会被拆成3个小数）
        # 时间数字特征：<=59的整数，且连续3个出现
        time_indices = set()
        for i in range(len(numbers) - 2):
            if (numbers[i] <= 59 and numbers[i] == int(numbers[i]) and
                numbers[i+1] <= 59 and numbers[i+1] == int(numbers[i+1]) and
                numbers[i+2] <= 59 and numbers[i+2] == int(numbers[i+2])):
                time_indices.update([i, i+1, i+2])
        clean_numbers = [n for i, n in enumerate(numbers) if i not in time_indices]

        if len(clean_numbers) < 2:
            # 有效数字太少（可能只有残缺信息），跳过此行
            continue

        # 识别成交金额（最大值）
        amount = max(clean_numbers)

        # 识别成交量（整数且>=100）
        volume = None
        for n in sorted(clean_numbers):
            if n >= 100 and n == int(n) and n != amount:
                volume = int(n)
                break

        # 识别成交价：金额之外的、合理范围内的数字
        # 价格应满足：大于0，小于金额，不等于已识别的数量
        price = None
        price_candidates = [n for n in clean_numbers if n != amount and n != volume and n > 0 and n < amount]
        if price_candidates:
            # 优先选最接近 amount/100 的值（假设100股为常见手数）
            if volume:
                target = amount / volume
                price = min(price_candidates, key=lambda n: abs(n - target))
            else:
                # 无volume时，选金额之外最大的非整数数字（金额通常是整数，价格有小数）
                decimal_prices = [n for n in price_candidates if n != int(n)]
                if decimal_prices:
                    price = max(decimal_prices)
                else:
                    price = max(price_candidates)

        # 如果有金额和价格，用它们反推/修正成交量
        if amount and price and price > 0 and price < amount:
            calc_volume = round(amount / price)
            if calc_volume >= 1:
                if volume is None:
                    # 成交量丢失，用金额/价格反推
                    volume = calc_volume
                elif abs(volume - calc_volume) > max(1, calc_volume * 0.05):
                    # 成交量偏差超过5%，说明OCR识别不准，用反推值修正
                    # 例如 OCR把100识别成96，但金额/价格反推是100
                    volume = calc_volume

        # 仍无成交量或反推成交量不合理时跳过
        # 如果金额/价格反推出极小值（<10股），说明这条记录的金额或价格OCR识别有误
        if volume is None or (not any(n >= 100 and n == int(n) and n != amount for n in clean_numbers) and volume < 10):
            continue

        raw_records.append({
            '证券代码': stock_code,
            '证券名称': stock_name,
            '买卖类别': direction,
            '成交数量': volume,
            '成交价格': price,
            '成交金额': amount,
            'y': sum(r['y'] for r in row) / len(row)  # 记录行y坐标用于调试
        })

    # 第三步：按股票代码聚合，合并同一股票的名称
    # 用出现次数最多的名称作为该股票的正式名称
    from collections import Counter

    code_names = {}
    for rec in raw_records:
        code = rec['证券代码']
        name = rec['证券名称']
        if name and name != '未知':
            if code not in code_names:
                code_names[code] = []
            code_names[code].append(name)

    # 为每个代码选择最佳名称
    best_names = {}
    for code, names in code_names.items():
        counter = Counter(names)
        best_names[code] = counter.most_common(1)[0][0]

    # 填充缺失的名称
    for rec in raw_records:
        code = rec['证券代码']
        if not rec['证券名称'] or rec['证券名称'] == '未知':
            if code in best_names:
                rec['证券名称'] = best_names[code]
            else:
                rec['证券名称'] = '未知'
        # 用名称修正表覆盖OCR截断/误读的名称
        if code in STOCK_NAME_CORRECTIONS:
            rec['证券名称'] = STOCK_NAME_CORRECTIONS[code]

    # OCR数量修正 v4.5.2：用"数量×价格≈金额"来验证OCR读数准确性
    # 如果OCR读的数量不是100整数倍，且用最近100整数倍算出的金额更接近OCR读的金额，则修正
    for code in set(r['证券代码'] for r in raw_records):
        for direction in ['证券买入', '证券卖出']:
            recs = [r for r in raw_records if r['证券代码'] == code and r['买卖类别'] == direction]
            if len(recs) == 0:
                continue

            for rec in recs:
                qty = rec['成交数量']
                price = rec['成交价格']
                amount = rec['成交金额']
                if qty % 100 != 0 and qty > 0 and price and amount:
                    nearest_100 = round(qty / 100) * 100
                    if nearest_100 <= 0:
                        continue
                    # 计算两种假设下的"金额误差"
                    error_current = abs(qty * price - amount)
                    error_corrected = abs(nearest_100 * price - amount)
                    # 如果修正后误差更小，采纳修正值
                    if error_corrected < error_current * 0.5:  # 修正后误差不到原来的一半
                        rec['成交数量'] = nearest_100
                        rec['成交金额'] = round(price * nearest_100, 2)
                    elif abs(nearest_100 - qty) / qty < 0.10:  # 偏差<10%且修正后金额合理
                        rec['成交数量'] = nearest_100
                        rec['成交金额'] = round(price * nearest_100, 2)

    # 去重：同一y位置+同一方向+同一代码+同一数量+同一金额视为OCR重复读取
    # y位置用于区分同一股票多笔逐笔成交（同价同量不同行）与OCR重复读取（同行同量）
    seen = set()
    unique_records = []
    for rec in raw_records:
        key = (rec['证券代码'], rec['买卖类别'], rec['成交数量'], rec['成交金额'], round(rec.get('y', 0), 1))
        if key not in seen:
            seen.add(key)
            unique_records.append({
                '证券代码': rec['证券代码'],
                '证券名称': rec['证券名称'],
                '买卖类别': rec['买卖类别'],
                '成交类型': '成交',
                '成交数量': rec['成交数量'],
                '成交价格': rec['成交价格'],
                '成交金额': rec['成交金额']
            })

    df = pd.DataFrame(unique_records)
    print(f"  识别到 {len(df)} 条交易记录")
    return df


def process_image_file(image_path):
    """处理单个图片文件（根据来源路由到不同解析器）"""
    trading_date = extract_date_from_filename(image_path)
    source = get_source_from_filename(os.path.basename(image_path))

    print(f"\n{'='*80}")
    print(f"处理图片：{os.path.basename(image_path)}")
    print(f"交易日期：{trading_date}")
    print(f"数据来源：{source}")
    print('='*80)

    if source == '平安账户':
        df = parse_pingan_image_trades(image_path)
    else:
        df = parse_image_trades(image_path)

    if len(df) == 0:
        print("未识别到有效交易记录")
        return pd.DataFrame(), trading_date, 0, source

    buy_records = df[df['买卖类别'].str.contains('证券买入', na=False)].copy()
    sell_records = df[df['买卖类别'].str.contains('证券卖出', na=False)].copy()

    print(f"买入记录数：{len(buy_records)}")
    print(f"卖出记录数：{len(sell_records)}")

    profit_results = calculate_profits(df, buy_records, sell_records, trading_date, source)

    total_profit = sum(r['盈亏金额'] for r in profit_results)
    print(f"\n总盈亏：{total_profit:.2f} 元")
    print('='*80)

    result_df = pd.DataFrame(profit_results)
    return result_df, trading_date, total_profit, source


# ==================== 平安证券截图OCR解析 ====================

def _infer_code_by_name(partial_code, stock_name):
    """通过股票名称和部分代码推断完整代码
    当OCR只识别出5位或不完整代码时，尝试修正
    """
    if not stock_name or stock_name == '未知':
        return None

    # 如果同一批次中已经有其他行识别出了完整代码+相同名称，优先匹配
    # （在parse_pingan_image_trades中通过known_codes参数传递）

    # 常见科创板代码修正：68开头缺位的
    digits = re.sub(r'[^0-9]', '', partial_code)
    if len(digits) == 5 and (digits.startswith('68') or digits.startswith('63')):
        # 尝试在688xxx范围内搜索已知股票
        pass  # 由外层调用者通过已知代码匹配

    return None


# 已知股票代码映射（用于OCR识别不完整时通过名称匹配）
KNOWN_STOCK_NAMES = {
    '厦钨新能': '688778', '厦钨新能源': '688778', '度钨新能源': '688778',
    '辰钨新能源': '688778', '钨新能源': '688778',
}


def _lookup_code_by_name(name):
    """通过模糊名称匹配查找股票代码"""
    if not name or name == '未知':
        return None
    # 精确匹配
    if name in KNOWN_STOCK_NAMES:
        return KNOWN_STOCK_NAMES[name]
    # 模糊匹配：名称包含关键词
    for key, code in KNOWN_STOCK_NAMES.items():
        if len(key) >= 3 and key[:3] in name:
            return code
        if len(name) >= 3 and name[:3] in key:
            return code
    return None


def parse_pingan_image_trades(image_path):
    """解析平安证券成交明细截图
    平安截图特点：严格的列式表格，有固定表头
    列：成交时间 | 证券代码 | 证券名称 | 操作 | 成交量 | 成交均价 | 成交金额
    """
    # 修复Windows下torch/getpass环境变量问题
    import os as _os
    if not _os.environ.get('USERNAME'):
        _os.environ['USERNAME'] = 'Ryan'
    if not _os.environ.get('USER'):
        _os.environ['USER'] = 'Ryan'

    try:
        import easyocr
    except ImportError:
        print(f"[错误] 缺少OCR依赖，请运行：pip install easyocr")
        return pd.DataFrame()

    import numpy as np
    from PIL import Image

    print(f"  正在识别平安截图：{os.path.basename(image_path)}")
    reader = get_ocr_reader()

    # 使用PIL读取并通过numpy传给easyocr（避免OpenCV路径问题）
    img = Image.open(image_path)
    # 如果图片太小，放大3倍提高OCR识别率
    w, h = img.size
    scale = 1
    if h < 300:
        scale = 3
        img = img.resize((w * scale, h * scale), Image.LANCZOS)
        print(f"  图片已放大 ({w}x{h} → {w*scale}x{h*scale})")
    img_array = np.array(img)
    result = reader.readtext(img_array)

    # 提取所有文本项，保留位置信息（还原坐标到原始比例）
    items = []
    for bbox, text, conf in result:
        y_center = (bbox[0][1] + bbox[2][1]) / 2 / scale
        x_center = (bbox[0][0] + bbox[2][0]) / 2 / scale
        items.append({'text': text.strip(), 'y': y_center, 'x': x_center, 'conf': conf})

    items.sort(key=lambda r: r['y'])

    # 第一步：识别表头行，确定各列的x范围
    # 表头关键词映射（OCR可能误读，允许多个匹配模式）
    header_patterns = {
        '成交时间': ['成交时间'],
        '证券代码': ['证券代码', '代码'],
        '证券名称': ['证券名称', '名称'],
        '操作': ['操作'],
        '成交量': ['成交量', '成交粼', '成交数量', '数量'],
        '成交均价': ['成交均价', '均价'],
        '成交金额': ['成交金额', '金额', '成交额'],
    }

    header_keywords = {k: None for k in header_patterns}
    header_y = None
    used_positions = set()

    for item in items:
        text = item['text']
        # 先精确匹配
        for col_name, patterns in header_patterns.items():
            if header_keywords[col_name] is not None:
                continue  # 已匹配
            for pat in patterns:
                if pat in text:
                    header_keywords[col_name] = item['x']
                    header_y = item['y']
                    used_positions.add(item['x'])
                    break

    # 如果某些列没识别到表头，用已识别的列推算
    found_headers = {k: v for k, v in header_keywords.items() if v is not None}
    if len(found_headers) < 3:
        print("  [警告] 未能识别足够表头列，尝试回退到手机截图解析模式")
        return parse_image_trades(image_path)  # 回退到通用解析

    # 确定列边界（每列取两个相邻表头x的中点作为分界）
    sorted_headers = sorted(found_headers.items(), key=lambda kv: kv[1])
    column_ranges = {}
    for i, (col_name, col_x) in enumerate(sorted_headers):
        if i == 0:
            left = 0
        else:
            left = (sorted_headers[i-1][1] + col_x) / 2
        if i == len(sorted_headers) - 1:
            right = 99999
        else:
            right = (col_x + sorted_headers[i+1][1]) / 2
        column_ranges[col_name] = (left, right)

    print(f"  识别到表头列：{list(column_ranges.keys())}")

    # 第二步：按y坐标分组成数据行（跳过表头行）
    data_items = [item for item in items if item['y'] > header_y + 5]
    if not data_items:
        print("  未找到数据行")
        return pd.DataFrame()

    rows = []
    current_row = []
    last_y = None
    y_threshold = 15  # 平安截图行间距较小

    for item in data_items:
        if last_y is None or abs(item['y'] - last_y) <= y_threshold:
            current_row.append(item)
        else:
            if current_row:
                rows.append(current_row)
            current_row = [item]
        last_y = item['y']

    if current_row:
        rows.append(current_row)

    # 第三步：从每行提取各列数据
    raw_records = []
    for row in rows:
        record = {}
        for col_name, (left, right) in column_ranges.items():
            col_texts = [r['text'] for r in row if left <= r['x'] < right]
            if col_texts:
                # 取该范围内置信度最高的文本
                best = max(
                    [r for r in row if left <= r['x'] < right],
                    key=lambda r: r['conf'],
                    default=None
                )
                record[col_name] = best['text'] if best else ' '.join(col_texts)
            else:
                record[col_name] = ''

        raw_records.append(record)

    # 第四步：数据清洗和转换
    trades = []
    for rec in raw_records:
        # 股票代码清洗：修正OCR常见误读
        code = rec.get('证券代码', '').strip()
        # 去除非数字字符
        code_digits = re.sub(r'[^0-9]', '', code)

        # 修正常见OCR错误模式
        if code_digits.startswith('633') and len(code_digits) == 6:
            code_digits = '688' + code_digits[3:]  # 633xxx → 688xxx
        elif code_digits.startswith('655') and len(code_digits) == 6:
            code_digits = '688' + code_digits[3:]  # 655xxx → 688xxx

        # 修正缺位：如 68775 → 688778（OCR漏了数字）
        if len(code_digits) == 5:
            code_match = re.search(r'(\d{6})', code)
            if not code_match:
                # 尝试常见的科创板补位模式
                if code_digits.startswith('688') or code_digits.startswith('68'):
                    # 可能漏了中间的数字，用已知代码表推断
                    pass  # 无法可靠推断，保留原始

        # 提取6位数字
        code_match = re.search(r'(\d{6})', code_digits if len(code_digits) >= 6 else code)
        if code_match:
            code = code_match.group(1)
        elif len(code_digits) == 6:
            code = code_digits
        else:
            # 代码不完整，尝试通过已识别到的股票名称推断
            code = _infer_code_by_name(code, rec.get('证券名称', ''))
            if not code:
                # 通过已知名称映射查找
                code = _lookup_code_by_name(rec.get('证券名称', ''))
            if not code:
                print(f"    [跳过] 无法识别股票代码: {rec.get('证券代码', '')}")
                continue

        # 股票名称清洗
        name = rec.get('证券名称', '').strip()
        if not name:
            name = '未知'

        # 操作方向清洗
        operation = rec.get('操作', '').strip()
        # 修正常见OCR误读
        if any(kw in operation for kw in ['卖', '觌出', '卖出']):
            direction = '证券卖出'
        elif any(kw in operation for kw in ['买', '买入']):
            direction = '证券买入'
        else:
            print(f"    [跳过] 无法判断买卖方向: {operation}")
            continue

        # 成交量清洗
        volume_text = rec.get('成交量', '').strip()
        volume = None
        if volume_text:
            vol_clean = re.sub(r'[^0-9]', '', volume_text)
            try:
                volume = int(vol_clean)
            except ValueError:
                pass

        # 成交金额清洗
        amount_text = rec.get('成交金额', '').strip()
        amount = None
        if amount_text:
            amt_clean = re.sub(r'[^0-9.]', '', amount_text)
            # 处理多个小数点的情况
            parts = amt_clean.split('.')
            if len(parts) > 2:
                amt_clean = parts[0] + '.' + ''.join(parts[1:])
            try:
                amount = float(amt_clean)
            except ValueError:
                pass

        # 成交均价清洗
        price_text = rec.get('成交均价', '').strip()
        price = None
        if price_text:
            price_clean = re.sub(r'[^0-9.]', '', price_text)
            parts = price_clean.split('.')
            if len(parts) > 2:
                price_clean = parts[0] + '.' + ''.join(parts[1:])
            try:
                price = float(price_clean)
            except ValueError:
                pass

        # 交叉验证与修正：如果 amount 存在，用它反推更可靠的 price/volume
        if amount and volume and price:
            calculated = round(volume * price, 2)
            if abs(calculated - amount) > max(amount * 0.02, 1):
                # 差异超过2%或1元，说明price可能误读，用amount反推
                price = round(amount / volume, 4)
        elif amount and volume and not price:
            price = round(amount / volume, 4)
        elif amount and price and not volume:
            volume = int(round(amount / price))
        elif volume and price and not amount:
            amount = round(volume * price, 2)

        # 跳过无效记录
        if not code or not name or volume is None:
            continue

        trades.append({
            '证券代码': code,
            '证券名称': name,
            '买卖类别': direction,
            '成交类型': '成交',
            '成交数量': volume,
            '成交价格': price or 0,
            '成交金额': amount or 0
        })

    # 去重：同代码+同方向+同数量+同金额视为重复
    seen = set()
    unique_records = []
    for t in trades:
        key = (t['证券代码'], t['买卖类别'], t['成交数量'], t['成交金额'])
        if key not in seen:
            seen.add(key)
            unique_records.append(t)

    df = pd.DataFrame(unique_records)
    print(f"  识别到 {len(df)} 条交易记录")
    if len(df) > 0:
        for _, row in df.iterrows():
            print(f"    {row['证券名称']}({row['证券代码']}) {row['买卖类别']} "
                  f"数量={row['成交数量']} 均价={row['成交价格']:.4f} 金额={row['成交金额']:.2f}")
    return df


# ==================== Excel处理 ====================

def parse_pingan_excel(file_path):
    """解析平安证券导出的成交记录
    平安导出文件特点：
    - 文件名含"平安"，扩展名可能是 .xls 但实际上可能是 TSV 格式（GBK编码）
    - 也可能是真正的 .xls 老格式
    - 第一行为列名，数据从第二行开始
    - 列：成交时间 | 证券代码 | 证券名称 | 操作 | 成交数量 | 成交均价 | 成交金额 | ...
    - 证券代码用格式：="688778"（需要剥离）
    - 操作用"买入"/"卖出"（需映射为标准格式）
    """
    # 先尝试作为TSV读取（GBK编码，平安常见导出格式）
    try:
        raw_df = pd.read_csv(file_path, sep='\t', encoding='gbk', dtype=str)
    except Exception:
        # 回退到xlrd引擎读真正的.xls
        engine = 'xlrd' if file_path.lower().endswith('.xls') else None
        raw_df = pd.read_excel(file_path, engine=engine)

    # 只取需要的列，映射到标准字段名
    col_map = {
        '证券代码': '证券代码', '证券名称': '证券名称',
        '操作': '买卖类别', '成交数量': '成交数量',
        '成交均价': '成交价格', '成交金额': '成交金额',
    }

    df = pd.DataFrame()
    name_col = None
    for src_col, dst_col in col_map.items():
        if src_col in raw_df.columns:
            val = raw_df[src_col]
            if dst_col == '证券代码':
                # 清洗 ="688778" 格式
                val = val.astype(str).str.replace('=', '').str.replace('"', '').str.strip()
            df[dst_col] = val
        elif dst_col == '证券名称' and '证券名称' in raw_df.columns:
            name_col = '证券名称'
            df['证券名称'] = raw_df['证券名称']

    # 添加统一的成交类型
    df['成交类型'] = '成交'

    # 映射操作字段：买入→证券买入，卖出→证券卖出
    df['买卖类别'] = df['买卖类别'].apply(
        lambda x: '证券买入' if '买' in str(x) else ('证券卖出' if '卖' in str(x) else str(x))
    )

    # 证券代码清洗（去除非数字字符）
    df['证券代码'] = df['证券代码'].astype(str).str.replace(r'[^0-9]', '', regex=True)
    df['成交数量'] = pd.to_numeric(df['成交数量'], errors='coerce')
    df['成交价格'] = pd.to_numeric(df['成交价格'], errors='coerce')
    df['成交金额'] = pd.to_numeric(df['成交金额'], errors='coerce')

    # 去空行，验证代码6位
    df = df.dropna(subset=['证券代码'])
    df = df[df['证券代码'].str.len() == 6]

    print(f"  识别到 {len(df)} 条交易记录")
    if len(df) > 0:
        name_col_ref = name_col if name_col and name_col in raw_df.columns else None
        for i, row in df.iterrows():
            name = raw_df.loc[i, name_col_ref] if name_col_ref and i in raw_df.index else row.get('证券名称', '')
            print(f"    {name}({row['证券代码']}) {row['买卖类别']} "
                  f"数量={int(row['成交数量'])} 均价={row['成交价格']:.4f} 金额={row['成交金额']:.2f}")

    return df


def process_excel_file(input_file):
    """处理单个Excel文件（根据来源路由到不同解析器）"""
    trading_date = extract_date_from_filename(input_file)
    source = get_source_from_filename(os.path.basename(input_file))

    print(f"\n{'='*80}")
    print(f"处理文件：{os.path.basename(input_file)}")
    print(f"交易日期：{trading_date}")
    print(f"数据来源：{source}")
    print('='*80)

    if source == '平安账户':
        df = parse_pingan_excel(input_file)
    else:
        df = pd.read_excel(input_file, skiprows=4, header=0)
        df.columns = ['证券代码', '证券名称', '买卖类别', '成交类型', '成交数量', '成交价格', '成交金额']
        df['证券代码'] = df['证券代码'].astype(str).str.replace('\t', '')
        df = df.dropna(subset=['证券代码'])
        df = df[df['证券代码'] != '证券代码']
        df['成交数量'] = pd.to_numeric(df['成交数量'], errors='coerce')
        df['成交价格'] = pd.to_numeric(df['成交价格'], errors='coerce')
        df['成交金额'] = pd.to_numeric(df['成交金额'], errors='coerce')

    buy_records = df[df['买卖类别'].str.contains('证券买入', na=False)].copy()
    sell_records = df[df['买卖类别'].str.contains('证券卖出', na=False)].copy()

    print(f"买入记录数：{len(buy_records)}")
    print(f"卖出记录数：{len(sell_records)}")

    profit_results = calculate_profits(df, buy_records, sell_records, trading_date, source)

    total_profit = sum(r['盈亏金额'] for r in profit_results)
    print(f"\n总盈亏：{total_profit:.2f} 元")
    print('='*80)

    result_df = pd.DataFrame(profit_results)
    return result_df, trading_date, total_profit, source


def validate_trades(df, source_tag=''):
    """数据校验：检查交易记录合理性，打印告警但不阻断流程
    返回清洗后的df（剔除严重异常行）
    """
    if len(df) == 0:
        return df

    warnings = []
    drop_indices = []

    for idx, row in df.iterrows():
        code = str(row.get('证券代码', ''))
        name = str(row.get('证券名称', ''))
        qty = row.get('成交数量', 0)
        price = row.get('成交价格', 0)
        amount = row.get('成交金额', 0)
        direction = str(row.get('买卖类别', ''))

        # 1. 代码长度校验
        if len(code) != 6 or not code.isdigit():
            warnings.append(f"  [异常] {name}({code}) 代码格式不正确")
            drop_indices.append(idx)
            continue

        # 2. 数量校验
        if qty is None or qty <= 0:
            warnings.append(f"  [异常] {name}({code}) {direction} 数量={qty}，已剔除")
            drop_indices.append(idx)
            continue

        # 3. 价格校验
        if price is not None and (price <= 0 or price > 10000):
            warnings.append(f"  [告警] {name}({code}) {direction} 价格={price} 异常")

        # 4. 金额校验：金额 ≈ 数量 × 价格
        if price is not None and price > 0 and amount is not None and amount > 0:
            expected = qty * price
            if abs(expected - amount) / max(amount, 1) > 0.05:
                warnings.append(f"  [告警] {name}({code}) {direction} 金额={amount} 但 数量×价格={expected:.2f}，偏差>5%")

        # 5. 数量非100整数倍（A股主板100股/手，科创板200股起）
        if qty > 0 and qty % 100 != 0:
            warnings.append(f"  [提示] {name}({code}) {direction} 数量={qty} 非100整数倍（可能是科创板200股起 or OCR误读）")

    if warnings:
        prefix = f"[{source_tag}] " if source_tag else ""
        print(f"\n  {prefix}--- 数据校验告警 ({len(warnings)}条) ---")
        for w in warnings:
            print(w)

    # 剔除严重异常行
    if drop_indices:
        df = df.drop(index=drop_indices).reset_index(drop=True)
        print(f"  已剔除 {len(drop_indices)} 条异常记录")

    return df


def calculate_profits(df, buy_records, sell_records, trading_date, source):
    """计算盈亏（通用逻辑，适用于Excel和图片输入）
    支持跨账户同股配对：当source包含多个账户时，按证券代码合并买卖后统一匹配
    支持单边交易标记：无法配对的买入/卖出记录标记为"未平仓"
    """
    profit_results = []
    all_stocks = set(df['证券代码'].unique())

    for stock_code in all_stocks:
        stock_name = df[df['证券代码'] == stock_code]['证券名称'].iloc[0]

        buys = buy_records[buy_records['证券代码'] == stock_code]
        sells = sell_records[sell_records['证券代码'] == stock_code]

        # 收集涉及的账户来源
        involved_sources = set()
        if '数据来源' in buys.columns:
            involved_sources.update(buys['数据来源'].unique())
        if '数据来源' in sells.columns:
            involved_sources.update(sells['数据来源'].unique())
        if not involved_sources:
            involved_sources = {source}

        # 确定显示的来源标签
        if len(involved_sources) > 1:
            display_source = '跨账户(' + '+'.join(sorted(involved_sources)) + ')'
        else:
            display_source = source

        total_buy_qty = buys['成交数量'].sum()
        total_sell_qty = sells['成交数量'].sum()

        if total_buy_qty == 0 and total_sell_qty == 0:
            continue

        # 无配对的情况：记录单边交易（未平仓）
        if total_buy_qty == 0 or total_sell_qty == 0:
            side = '仅买入' if total_buy_qty > 0 else '仅卖出'
            qty = total_buy_qty if total_buy_qty > 0 else total_sell_qty
            amt = buys['成交金额'].sum() if total_buy_qty > 0 else sells['成交金额'].sum()
            avg_price = amt / qty if qty > 0 else 0

            # 计算单边交易成本（只有佣金，没有印花税因为没有卖出；只有卖出时收印花税+佣金）
            if total_buy_qty > 0:
                commission = max(amt * COMMISSION_RATE, MIN_COMMISSION)
                stamp_duty = 0
            else:
                commission = max(amt * COMMISSION_RATE, MIN_COMMISSION)
                stamp_duty = round(amt * STAMP_DUTY_RATE, 2)
            total_cost = round(commission + stamp_duty, 2)

            profit_results.append({
                '日期': trading_date,
                '数据来源': display_source,
                '证券代码': stock_code,
                '证券名称': stock_name,
                '买入数量': int(total_buy_qty),
                '卖出数量': int(total_sell_qty),
                '匹配数量': 0,
                '买入均价': round(avg_price, 4) if total_buy_qty > 0 else 0,
                '卖出均价': round(avg_price, 4) if total_sell_qty > 0 else 0,
                '买入金额': round(buys['成交金额'].sum(), 2) if total_buy_qty > 0 else 0,
                '卖出金额': round(sells['成交金额'].sum(), 2) if total_sell_qty > 0 else 0,
                '毛盈亏': 0,
                '佣金': commission,
                '印花税': stamp_duty,
                '交易成本': total_cost,
                '盈亏金额': round(-total_cost, 2),
                '盈亏比例': f"⚠️{side}未平仓"
            })
            print(f"股票：{stock_name} ({stock_code}) → {side}未平仓，数量={int(qty)}")
            continue

        matched_qty = min(total_buy_qty, total_sell_qty)

        if matched_qty == 0:
            continue

        total_buy_amt = buys['成交金额'].sum()
        avg_buy_price = total_buy_amt / total_buy_qty
        matched_buy_amt = avg_buy_price * matched_qty

        total_sell_amt = sells['成交金额'].sum()
        avg_sell_price = total_sell_amt / total_sell_qty
        matched_sell_amt = avg_sell_price * matched_qty

        profit = matched_sell_amt - matched_buy_amt  # 毛盈亏

        # 交易成本计算
        buy_commission = max(matched_buy_amt * COMMISSION_RATE, MIN_COMMISSION)
        sell_commission = max(matched_sell_amt * COMMISSION_RATE, MIN_COMMISSION)
        commission = round(buy_commission + sell_commission, 2)
        stamp_duty = round(matched_sell_amt * STAMP_DUTY_RATE, 2)
        total_cost = round(commission + stamp_duty, 2)
        net_profit = round(profit - total_cost, 2)
        profit_pct = (net_profit / matched_buy_amt) * 100 if matched_buy_amt != 0 else 0

        # 如果有未匹配的部分，也记录
        unmatched_buy_qty = total_buy_qty - matched_qty
        unmatched_sell_qty = total_sell_qty - matched_qty

        profit_results.append({
            '日期': trading_date,
            '数据来源': display_source,
            '证券代码': stock_code,
            '证券名称': stock_name,
            '买入数量': int(total_buy_qty),
            '卖出数量': int(total_sell_qty),
            '匹配数量': int(matched_qty),
            '买入均价': round(avg_buy_price, 4),
            '卖出均价': round(avg_sell_price, 4),
            '买入金额': round(matched_buy_amt, 2),
            '卖出金额': round(matched_sell_amt, 2),
            '毛盈亏': round(profit, 2),
            '佣金': commission,
            '印花税': stamp_duty,
            '交易成本': total_cost,
            '盈亏金额': net_profit,
            '盈亏比例': f"{profit_pct:.2f}%"
        })

        print(f"股票：{stock_name} ({stock_code})" + (f" [跨账户]" if len(involved_sources) > 1 else ""))
        print(f"  买入：数量={total_buy_qty:.0f}, 均价={avg_buy_price:.4f}, 金额={matched_buy_amt:.2f}")
        print(f"  卖出：数量={total_sell_qty:.0f}, 均价={avg_sell_price:.4f}, 金额={matched_sell_amt:.2f}")
        print(f"  毛盈亏：{profit:.2f} 元")
        print(f"  交易成本：佣金={commission:.2f}, 印花税={stamp_duty:.2f}, 合计={total_cost:.2f}")
        print(f"  净盈亏：{net_profit:.2f} 元 ({profit_pct:.2f}%)")

        # 记录未匹配部分
        if unmatched_buy_qty > 0:
            unmatched_buy_amt = avg_buy_price * unmatched_buy_qty
            uc = max(unmatched_buy_amt * COMMISSION_RATE, MIN_COMMISSION)
            profit_results.append({
                '日期': trading_date,
                '数据来源': display_source,
                '证券代码': stock_code,
                '证券名称': stock_name,
                '买入数量': int(unmatched_buy_qty),
                '卖出数量': 0,
                '匹配数量': 0,
                '买入均价': round(avg_buy_price, 4),
                '卖出均价': 0,
                '买入金额': round(unmatched_buy_amt, 2),
                '卖出金额': 0,
                '毛盈亏': 0,
                '佣金': round(uc, 2),
                '印花税': 0,
                '交易成本': round(uc, 2),
                '盈亏金额': round(-uc, 2),
                '盈亏比例': "⚠️多买未平仓"
            })
            print(f"  ⚠️ 多买入 {int(unmatched_buy_qty)} 股未平仓")

        if unmatched_sell_qty > 0:
            unmatched_sell_amt = avg_sell_price * unmatched_sell_qty
            uc = max(unmatched_sell_amt * COMMISSION_RATE, MIN_COMMISSION)
            usd = round(unmatched_sell_amt * STAMP_DUTY_RATE, 2)
            profit_results.append({
                '日期': trading_date,
                '数据来源': display_source,
                '证券代码': stock_code,
                '证券名称': stock_name,
                '买入数量': 0,
                '卖出数量': int(unmatched_sell_qty),
                '匹配数量': 0,
                '买入均价': 0,
                '卖出均价': round(avg_sell_price, 4),
                '买入金额': 0,
                '卖出金额': round(unmatched_sell_amt, 2),
                '毛盈亏': 0,
                '佣金': round(uc, 2),
                '印花税': usd,
                '交易成本': round(uc + usd, 2),
                '盈亏金额': round(-(uc + usd), 2),
                '盈亏比例': "⚠️多卖未平仓"
            })
            print(f"  ⚠️ 多卖出 {int(unmatched_sell_qty)} 股未平仓")

    return profit_results


# ==================== Excel汇总 ====================

def append_to_excel(result_df, trading_date, source):
    """追加数据到Excel汇总文件（支持按日期去重，同一天的数据整体替换）"""
    if len(result_df) == 0:
        return

    if os.path.exists(EXCEL_OUTPUT):
        existing_df = pd.read_excel(EXCEL_OUTPUT)
        # 删除该日期的旧数据（整日替换，因为跨账户匹配可能改变所有记录）
        existing_df = existing_df[existing_df['日期'] != trading_date]
        combined_df = pd.concat([existing_df, result_df], ignore_index=True)
        combined_df = combined_df.sort_values(['日期', '数据来源']).reset_index(drop=True)
    else:
        combined_df = result_df

    wb = Workbook()
    ws = wb.active
    ws.title = "股票盈亏汇总"

    headers = ['日期', '数据来源', '证券代码', '证券名称', '买入数量', '卖出数量', '匹配数量',
               '买入均价', '卖出均价', '买入金额', '卖出金额',
               '毛盈亏', '佣金', '印花税', '交易成本', '盈亏金额', '盈亏比例']

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, row in enumerate(combined_df.itertuples(index=False), 2):
        ws.cell(row=idx, column=1, value=row.日期)
        ws.cell(row=idx, column=2, value=row.数据来源)
        ws.cell(row=idx, column=3, value=row.证券代码)
        ws.cell(row=idx, column=4, value=row.证券名称)
        ws.cell(row=idx, column=5, value=row.买入数量)
        ws.cell(row=idx, column=6, value=row.卖出数量)
        ws.cell(row=idx, column=7, value=row.匹配数量)
        ws.cell(row=idx, column=8, value=row.买入均价)
        ws.cell(row=idx, column=9, value=row.卖出均价)
        ws.cell(row=idx, column=10, value=row.买入金额)
        ws.cell(row=idx, column=11, value=row.卖出金额)

        # 毛盈亏
        gross_cell = ws.cell(row=idx, column=12, value=row.毛盈亏)
        if row.毛盈亏 > 0:
            gross_cell.font = Font(color="FF0000", bold=True)
        else:
            gross_cell.font = Font(color="00B050", bold=True)

        ws.cell(row=idx, column=13, value=row.佣金)
        ws.cell(row=idx, column=14, value=row.印花税)
        ws.cell(row=idx, column=15, value=row.交易成本)

        profit_cell = ws.cell(row=idx, column=16, value=row.盈亏金额)
        if row.盈亏金额 > 0:
            profit_cell.font = Font(color="FF0000", bold=True)
        else:
            profit_cell.font = Font(color="00B050", bold=True)

        ws.cell(row=idx, column=17, value=row.盈亏比例)

        for col in range(1, 18):
            ws.cell(row=idx, column=col).alignment = Alignment(horizontal='center', vertical='center')

    column_widths = [12, 12, 12, 14, 10, 10, 10, 12, 12, 14, 14, 14, 12, 12, 12, 14, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    action = "更新" if os.path.exists(EXCEL_OUTPUT) else "创建"
    wb.save(EXCEL_OUTPUT)
    print(f"Excel汇总已{action}（按日期去重，整日替换）：{EXCEL_OUTPUT}")


# ==================== HTML报告生成 ====================

def generate_html_report_from_summary(trading_date):
    """从汇总文件中提取当日数据生成HTML报告"""
    html_filename = f"{trading_date}-股票交易盈亏报告.html"
    html_path = os.path.join(REPORTS_DIR, html_filename)

    if not os.path.exists(EXCEL_OUTPUT):
        print(f"汇总文件不存在，无法生成 {trading_date} 的报告")
        return

    summary_df = pd.read_excel(EXCEL_OUTPUT)
    day_df = summary_df[summary_df['日期'] == trading_date].copy()

    if len(day_df) == 0:
        print(f"汇总文件中未找到 {trading_date} 的数据")
        return

    total_profit = day_df['盈亏金额'].sum()
    total_commission = day_df['佣金'].sum() if '佣金' in day_df.columns else 0
    total_stamp_duty = day_df['印花税'].sum() if '印花税' in day_df.columns else 0
    total_cost = total_commission + total_stamp_duty

    # 按数据来源分组统计
    sources = day_df['数据来源'].unique()
    source_summary = ""
    for src in sources:
        src_df = day_df[day_df['数据来源'] == src]
        src_profit = src_df['盈亏金额'].sum()
        source_summary += f"<span class='source-tag source-{src[:2]}'>{src}: {'+' if src_profit > 0 else ''}¥{src_profit:,.2f}</span>"

    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>股票交易盈亏报告 - {trading_date}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Microsoft YaHei', 'SimHei', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{ font-size: 32px; margin-bottom: 10px; font-weight: 600; }}
        .header .date {{ font-size: 18px; opacity: 0.9; margin-bottom: 8px; }}
        .header .sources {{ font-size: 14px; opacity: 0.85; }}
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        .card {{
            background: white;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            text-align: center;
            transition: transform 0.3s;
        }}
        .card:hover {{ transform: translateY(-5px); }}
        .card .title {{ color: #666; font-size: 14px; margin-bottom: 10px; text-transform: uppercase; letter-spacing: 1px; }}
        .card .value {{ font-size: 32px; font-weight: bold; margin-bottom: 5px; }}
        .card .value.profit {{ color: #e74c3c; }}
        .card .value.loss {{ color: #27ae60; }}
        .card .value.neutral {{ color: #95a5a6; }}
        .table-container {{ padding: 30px; }}
        .table-title {{
            font-size: 24px;
            color: #2c3e50;
            margin-bottom: 20px;
            padding-left: 15px;
            border-left: 4px solid #667eea;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
            table-layout: fixed;
        }}
        thead {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }}
        th {{ color: white; padding: 15px 8px; text-align: center; font-weight: 600; font-size: 14px; }}
        td {{ padding: 15px 8px; text-align: center; border-bottom: 1px solid #ecf0f1; font-size: 14px; }}
        tbody tr:hover {{ background: #f8f9fa; }}
        .profit-amount {{ color: #e74c3c; font-weight: bold; }}
        .loss-amount {{ color: #27ae60; font-weight: bold; }}
        .footer {{
            padding: 20px;
            text-align: center;
            color: #7f8c8d;
            font-size: 14px;
            background: #f8f9fa;
            border-top: 1px solid #ecf0f1;
        }}
        .no-data {{ text-align: center; padding: 40px; color: #7f8c8d; }}
        .source-tag {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 12px;
            margin: 0 4px;
        }}
        .source-两融 {{ background: #e3f2fd; color: #1976d2; }}
        .source-手机 {{ background: #fff3e0; color: #f57c00; }}
        .source-平安 {{ background: #e8f5e9; color: #388e3c; }}
        /* 列宽设置 */
        th:nth-child(1), td:nth-child(1) {{ width: 10%; }} /* 数据来源 */
        th:nth-child(2), td:nth-child(2) {{ width: 14%; }} /* 证券名称 */
        th:nth-child(3), td:nth-child(3) {{ width: 8%; }} /* 买入数量 */
        th:nth-child(4), td:nth-child(4) {{ width: 8%; }} /* 卖出数量 */
        th:nth-child(5), td:nth-child(5) {{ width: 8%; }} /* 匹配数量 */
        th:nth-child(6), td:nth-child(6) {{ width: 10%; }} /* 买入均价 */
        th:nth-child(7), td:nth-child(7) {{ width: 10%; }} /* 卖出均价 */
        th:nth-child(8), td:nth-child(8) {{ width: 11%; }} /* 买入金额 */
        th:nth-child(9), td:nth-child(9) {{ width: 11%; }} /* 卖出金额 */
        th:nth-child(10), td:nth-child(10) {{ width: 8%; }} /* 佣金 */
        th:nth-child(11), td:nth-child(11) {{ width: 8%; }} /* 印花税 */
        th:nth-child(12), td:nth-child(12) {{ width: 12%; }} /* 盈亏金额 */
        th:nth-child(13), td:nth-child(13) {{ width: 10%; }} /* 盈亏比例 */
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📈 股票交易盈亏报告</h1>
            <div class="date">交易日期：{trading_date}</div>
            <div class="sources">{source_summary}</div>
        </div>
        <div class="summary-cards">
            <div class="card">
                <div class="title">交易股票数</div>
                <div class="value neutral">{len(day_df)}</div>
            </div>
            <div class="card">
                <div class="title">总买入金额</div>
                <div class="value neutral">¥{day_df['买入金额'].sum():,.2f}</div>
            </div>
            <div class="card">
                <div class="title">总卖出金额</div>
                <div class="value neutral">¥{day_df['卖出金额'].sum():,.2f}</div>
            </div>
            <div class="card">
                <div class="title">佣金合计</div>
                <div class="value neutral">¥{total_commission:,.2f}</div>
            </div>
            <div class="card">
                <div class="title">印花税合计</div>
                <div class="value neutral">¥{total_stamp_duty:,.2f}</div>
            </div>
            <div class="card">
                <div class="title">交易成本合计</div>
                <div class="value neutral">¥{total_cost:,.2f}</div>
            </div>
            <div class="card">
                <div class="title">总盈亏金额</div>
                <div class="value {'profit' if total_profit > 0 else 'loss'}">
                    {'+' if total_profit > 0 else ''}¥{total_profit:,.2f}
                </div>
            </div>
        </div>
        <div class="table-container">
            <h2 class="table-title">交易明细</h2>
"""

    if len(day_df) > 0:
        html_content += """
            <table>
                <thead>
                    <tr>
                        <th>数据来源</th>
                        <th>证券名称</th>
                        <th>买入数量</th>
                        <th>卖出数量</th>
                        <th>匹配数量</th>
                        <th>买入均价</th>
                        <th>卖出均价</th>
                        <th>买入金额</th>
                        <th>卖出金额</th>
                        <th>佣金</th>
                        <th>印花税</th>
                        <th>盈亏金额</th>
                        <th>盈亏比例</th>
                    </tr>
                </thead>
                <tbody>
"""
        for _, row in day_df.iterrows():
            profit_class = 'profit-amount' if row['盈亏金额'] > 0 else 'loss-amount'
            profit_sign = '+' if row['盈亏金额'] > 0 else ''
            commission_val = row['佣金'] if '佣金' in row.index else 0
            stamp_val = row['印花税'] if '印花税' in row.index else 0
            html_content += f"""
                    <tr>
                        <td><span class="source-tag source-{row['数据来源'][:2]}">{row['数据来源']}</span></td>
                        <td><strong>{row['证券名称']}</strong></td>
                        <td>{row['买入数量']:,.0f}</td>
                        <td>{row['卖出数量']:,.0f}</td>
                        <td>{row['匹配数量']:,.0f}</td>
                        <td>¥{row['买入均价']:,.4f}</td>
                        <td>¥{row['卖出均价']:,.4f}</td>
                        <td>¥{row['买入金额']:,.2f}</td>
                        <td>¥{row['卖出金额']:,.2f}</td>
                        <td>¥{commission_val:,.2f}</td>
                        <td>¥{stamp_val:,.2f}</td>
                        <td class="{profit_class}">{profit_sign}¥{row['盈亏金额']:,.2f}</td>
                        <td class="{profit_class}">{row['盈亏比例']}</td>
                    </tr>
"""
        html_content += """
                </tbody>
            </table>
"""
    else:
        html_content += """
            <div class="no-data">
                <p>暂无匹配的买卖记录</p>
            </div>
"""

    html_content += f"""
        </div>
        <div class="footer">
            <p>报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML报告已生成（从汇总提取）：{html_path}")


# ==================== 汇总可视化报告（v3.0 交互式） ====================

def compute_monthly_cross():
    """复用跨天配对分析脚本，算每月系统现有/跨天释放/修正后，用于汇总报告月度趋势三柱图。
    返回 {ym: {'sys':..,'cross':..,'corrected':..}}。失败返回空字典，前端自动降级为单柱。"""
    try:
        import importlib
        kt = importlib.import_module('跨天配对分析')
        df = kt.load()
        out = {}
        for ym in sorted(df['ym'].unique()):
            r = kt.analyze(df[df['ym'] == ym])
            out[ym] = {'sys': r['sys_total'], 'cross': r['cross_net'], 'corrected': r['corrected']}
        return out
    except Exception as e:
        print(f"[warn] 月度趋势图跨天数据计算失败，降级为仅显示系统现有盈亏：{e}")
        return {}


def generate_summary_html():
    """生成交互式汇总可视化HTML报告"""
    if not os.path.exists(EXCEL_OUTPUT):
        print("未找到汇总文件，跳过汇总报告生成")
        return

    df = pd.read_excel(EXCEL_OUTPUT)
    if len(df) == 0:
        print("汇总文件无数据，跳过汇总报告生成")
        return

    df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d')

    import json
    records = []
    for _, row in df.iterrows():
        records.append({
            'date': str(row['日期']),
            'source': str(row['数据来源']),
            'code': str(row['证券代码']),
            'name': str(row['证券名称']),
            'buyQty': int(row['买入数量']),
            'sellQty': int(row['卖出数量']),
            'matchQty': int(row['匹配数量']),
            'buyPrice': round(float(row['买入均价']), 4),
            'sellPrice': round(float(row['卖出均价']), 4),
            'buyAmount': round(float(row['买入金额']), 2),
            'sellAmount': round(float(row['卖出金额']), 2),
            'commission': round(float(row['佣金']), 2) if '佣金' in row.index else 0,
            'stampDuty': round(float(row['印花税']), 2) if '印花税' in row.index else 0,
            'totalCost': round(float(row['交易成本']), 2) if '交易成本' in row.index else 0,
            'profit': round(float(row['盈亏金额']), 2),
            'profitPct': str(row['盈亏比例'])
        })

    data_json = json.dumps(records, ensure_ascii=False)
    now_str = datetime.now().strftime('%Y-%m-%d')
    gen_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    html_path = os.path.join(REPORTS_DIR, '汇总可视化报告.html')

    # 读取独立模板文件并注入数据（模板中用 {{__VAR__}} 标记占位符）
    if not os.path.exists(SUMMARY_TEMPLATE):
        print(f"[错误] 汇总报告模板不存在：{SUMMARY_TEMPLATE}")
        return
    with open(SUMMARY_TEMPLATE, 'r', encoding='utf-8') as f:
        html_content = f.read()
    html_content = html_content.replace('{__DATA_JSON__}', data_json)
    html_content = html_content.replace('{__NOW_STR__}', now_str)
    html_content = html_content.replace('{__GEN_TIME__}', gen_time)

    # 月度跨天三柱图数据（系统现有 / 跨天释放 / 修正后），与跨天配对脚本口径一致
    monthly_cross = compute_monthly_cross()
    html_content = html_content.replace('{__MONTHLY_CROSS__}', json.dumps(monthly_cross, ensure_ascii=False))

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"汇总可视化报告已生成：{html_path}")


# generate_html_template 已移除，HTML/CSS/JS 模板存放在 reports/templates/summary_report.html


# ==================== 辅助函数 ====================

def archive_file(input_file):
    """归档原始文件"""
    filename = os.path.basename(input_file)
    archive_path = os.path.join(HISTORY_DIR, filename)

    if os.path.exists(archive_path):
        base, ext = os.path.splitext(filename)
        timestamp = datetime.now().strftime('%H%M%S')
        filename = f"{base}_{timestamp}{ext}"
        archive_path = os.path.join(HISTORY_DIR, filename)

    shutil.move(input_file, archive_path)
    print(f"原始文件已归档：{archive_path}")


def find_input_files():
    """查找待处理的所有输入文件（Excel + 图片）"""
    excel_files = [f for f in glob.glob('*.xlsx') + glob.glob('*.xls')
                   if f not in ['股票交易盈亏汇总.xlsx']
                   and not f.startswith('~$')]

    image_extensions = ['*.png', '*.jpg', '*.jpeg']
    image_files = []
    for ext in image_extensions:
        image_files.extend(glob.glob(ext))

    return excel_files, image_files


# ==================== 主程序 ====================

def main():
    # 修复Windows终端emoji打印乱码/编码错误
    import io, sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    print("\n" + "="*80)
    print("股票交易盈亏分析系统 v4.5")
    print("支持输入：Excel文件（券商导出）+ 图片文件（手机App截图/平安证券截图）")
    print("="*80)

    excel_files, image_files = find_input_files()
    all_files = excel_files + image_files

    if not all_files:
        print("\n未找到待处理的文件")
        print("请将以下类型文件放到当前文件夹：")
        print("  - Excel文件：券商/平安导出的交易记录（如：2026-04-22-两融-当日成交汇总.xlsx 或 20260430_平安.xls）")
        print("  - 图片文件：手机App截图（如：2026-04-22-手机交易.png）")
        print("  - 图片文件：平安证券截图（如：20260430_平安.png）")
        generate_summary_html()
        return

    print(f"\n找到 {len(excel_files)} 个Excel文件，{len(image_files)} 个图片文件")

    # 按日期分组处理（同一天可能有Excel+图片两种输入）
    files_by_date = {}

    for f in excel_files:
        d = extract_date_from_filename(f)
        if d not in files_by_date:
            files_by_date[d] = {'excel': [], 'image': []}
        files_by_date[d]['excel'].append(f)

    for f in image_files:
        d = extract_date_from_filename(f)
        if d not in files_by_date:
            files_by_date[d] = {'excel': [], 'image': []}
        files_by_date[d]['image'].append(f)

    processed_dates = []

    # 按日期顺序处理
    for trading_date in sorted(files_by_date.keys()):
        date_files = files_by_date[trading_date]

        print(f"\n{'='*80}")
        print(f"处理日期：{trading_date}")
        print(f"  Excel文件：{len(date_files['excel'])} 个")
        print(f"  图片文件：{len(date_files['image'])} 个")
        print('='*80)

        # 收集该日期所有文件的买卖记录，用于跨账户合并匹配
        all_buy_records = []
        all_sell_records = []
        all_stock_info = {}  # code -> name
        file_sources = set()
        processed_files = []  # 记录已成功解析的文件，全部成功后再归档

        # 处理Excel文件：先解析，收集买卖记录
        for excel_file in date_files['excel']:
            try:
                source = get_source_from_filename(os.path.basename(excel_file))
                if source == '平安账户':
                    df = parse_pingan_excel(excel_file)
                else:
                    df = pd.read_excel(excel_file, skiprows=4, header=0)
                    df.columns = ['证券代码', '证券名称', '买卖类别', '成交类型', '成交数量', '成交价格', '成交金额']
                    df['证券代码'] = df['证券代码'].astype(str).str.replace('\t', '')
                    df = df.dropna(subset=['证券代码'])
                    df = df[df['证券代码'] != '证券代码']
                    df['成交数量'] = pd.to_numeric(df['成交数量'], errors='coerce')
                    df['成交价格'] = pd.to_numeric(df['成交价格'], errors='coerce')
                    df['成交金额'] = pd.to_numeric(df['成交金额'], errors='coerce')

                buy_records = df[df['买卖类别'].str.contains('证券买入', na=False)].copy()
                sell_records = df[df['买卖类别'].str.contains('证券卖出', na=False)].copy()

                # 标记来源
                buy_records['数据来源'] = source
                sell_records['数据来源'] = source
                file_sources.add(source)

                all_buy_records.append(buy_records)
                all_sell_records.append(sell_records)

                # 收集证券名称映射
                for _, row in df.iterrows():
                    code = str(row.get('证券代码', ''))
                    name = str(row.get('证券名称', ''))
                    if code and name and name != 'nan':
                        all_stock_info[code] = name

                print(f"  [{source}] Excel解析完成：买入{len(buy_records)}笔，卖出{len(sell_records)}笔")
                processed_files.append(excel_file)
            except Exception as e:
                print(f"\n[错误] 处理Excel文件 {excel_file} 时出错：{str(e)}")
                print(f"  文件未归档，可修正后重跑")
                continue

        # 处理图片文件：先解析，收集买卖记录
        for image_file in date_files['image']:
            try:
                source = get_source_from_filename(os.path.basename(image_file))
                if source == '平安账户':
                    df = parse_pingan_image_trades(image_file)
                else:
                    df = parse_image_trades(image_file)

                if len(df) == 0:
                    print(f"  [{source}] 图片未识别到有效交易记录")
                    continue

                buy_records = df[df['买卖类别'].str.contains('证券买入', na=False)].copy()
                sell_records = df[df['买卖类别'].str.contains('证券卖出', na=False)].copy()

                # 标记来源
                buy_records['数据来源'] = source
                sell_records['数据来源'] = source
                file_sources.add(source)

                all_buy_records.append(buy_records)
                all_sell_records.append(sell_records)

                # 收集证券名称映射
                for _, row in df.iterrows():
                    code = str(row.get('证券代码', ''))
                    name = str(row.get('证券名称', ''))
                    if code and name and name != 'nan':
                        all_stock_info[code] = name

                print(f"  [{source}] 图片解析完成：买入{len(buy_records)}笔，卖出{len(sell_records)}笔")
                processed_files.append(image_file)
            except Exception as e:
                print(f"\n[错误] 处理图片文件 {image_file} 时出错：{str(e)}")
                print(f"  文件未归档，可修正后重跑")
                continue

        # 合并所有买卖记录，统一跨账户匹配
        if all_buy_records or all_sell_records:
            merged_buys = pd.concat(all_buy_records, ignore_index=True) if all_buy_records else pd.DataFrame()
            merged_sells = pd.concat(all_sell_records, ignore_index=True) if all_sell_records else pd.DataFrame()

            # 构建统一的df用于calculate_profits
            all_dfs = []
            if len(merged_buys) > 0:
                all_dfs.append(merged_buys)
            if len(merged_sells) > 0:
                all_dfs.append(merged_sells)
            merged_df = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

            # 使用主来源标签（如果只有一个来源就用那个，否则用"多账户合并"）
            primary_source = '+'.join(sorted(file_sources)) if len(file_sources) > 1 else (list(file_sources)[0] if file_sources else '未知')

            print(f"\n  --- 跨账户合并匹配 ---")
            print(f"  合并买入：{len(merged_buys)} 笔")
            print(f"  合并卖出：{len(merged_sells)} 笔")
            print(f"  涉及来源：{', '.join(sorted(file_sources))}")

            # 数据校验：检查合理性，打印告警，剔除严重异常行
            merged_df = validate_trades(merged_df, primary_source)
            if len(merged_df) == 0:
                print("  [告警] 校验后无有效记录，跳过盈亏计算")
                profit_results = []
            else:
                # 重新拆分校验后的买卖记录
                merged_buys = merged_df[merged_df['买卖类别'].str.contains('证券买入', na=False)].copy()
                merged_sells = merged_df[merged_df['买卖类别'].str.contains('证券卖出', na=False)].copy()
                profit_results = calculate_profits(merged_df, merged_buys, merged_sells, trading_date, primary_source)

            if profit_results:
                result_df = pd.DataFrame(profit_results)
                total_profit = sum(r['盈亏金额'] for r in profit_results)
                print(f"\n  当日总盈亏：{total_profit:.2f} 元")
                append_to_excel(result_df, trading_date, primary_source)

        # 从汇总文件提取当日全部数据生成单日报告
        generate_html_report_from_summary(trading_date)
        processed_dates.append(trading_date)

        # 全部处理成功后才归档原始文件
        for f in processed_files:
            archive_file(f)

        print(f"\n日期 {trading_date} 处理完成")

    # 生成汇总可视化报告
    generate_summary_html()

    print("\n" + "="*80)
    print("[完成] 所有文件处理完成！")
    print("="*80)


if __name__ == "__main__":
    main()
